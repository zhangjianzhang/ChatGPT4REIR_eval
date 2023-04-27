#!/usr/bin/env python
# -*- coding: utf-8 -*-


import os
import re
import csv
import xlsxwriter
from openpyxl import load_workbook
from .stringUtils import jiema,encodeUTF8
import dill

def readExcelToList(excelFile):
	wb = load_workbook(excelFile)
	ws = wb.active
	numOfRow = ws.max_row
	numOfCol = ws.max_column
	table = []
	for row in range(1,numOfRow+1):
		rowList = []
		for col in range(1,numOfCol+1):
			cellValue = ws.cell(row=row, column=col).value
			rowList.append(cellValue)
		table.append(rowList)
	return table

def getFileList(dir, fileList):
	if os.path.isfile(dir):
		fileList.append(dir)
	elif os.path.isdir(dir):
		for s in os.listdir(dir):
			newDir = os.path.join(dir, s)
			getFileList(newDir, fileList)
	return fileList


def readStrFromFile(filePath):
	with open(filePath, "rb") as f:
		string = jiema(f.read().strip())
	return string


def readLinesFromFile(filePath):
	
	with open(filePath, "rb") as f:
		content = jiema(f.read().strip())
	lines = content.split("\n")
	return lines


def writeStrToFile(filePath, string):
	with open(filePath, "wb") as f:
		f.write(encodeUTF8(string))


def appendStrToFile(filePath, string):
	with open(filePath, "ab") as f:
		f.write(encodeUTF8(string))


def dumpToFile(filePath, content):
	with open(filePath, "wb") as f:
		dill.dump(content, f)


def loadFromFile(filePath):
	with open(filePath,'rb') as f:
		content = dill.load(f)
	return content


def loadJson(filePath):
	import json
	with open(filePath) as f:
		content = json.load(f)
	return content


def dumpJson(filePath,obj):
	import json
	content = json.dumps(obj, ensure_ascii=False, sort_keys=True, indent=4)
	writeStrToFile(filePath,content)

def appendToJson(filePath,obj):
	import json
	content = json.dumps(obj, ensure_ascii=False, sort_keys=True, indent=4)
	appendStrToFile(filePath,content)
	
def writeToExcel(excel, tableList):
	wb = xlsxwriter.Workbook(excel)
	ws = wb.add_worksheet('sheet1')
	for i, itemList in enumerate(tableList):
		newItemList = [jiema(item) if item else "" for item in itemList]
		for j, item in enumerate(newItemList):
			ws.write(i, j, item)
	wb.close()
	
def formatFilename(rawName):
	return re.sub('[\\\\/:*?"<>|\s]', '_', rawName)

def appendToCsv(filePath, table):
	with open(filePath,"a") as f:
		writer = csv.writer(f)
		for row in table:
			# newRow = [encodeUTF8(item) if item else "" for item in row]
			# writer.writerow(newRow)
			writer.writerow(row)
	

def readCsvToList(csvFile):
	table = []
	with open(csvFile) as f:
		csvReader = csv.reader(f)
		for row in csvReader:
			table.append(row)
	return table

def writeToCsv(filePath,table):
	with open(filePath,"w") as f:
		writer = csv.writer(f)
		for row in table:
			# newRow = [encodeUTF8(item) if item else "" for item in row]
			# writer.writerow(newRow)
			writer.writerow(row)

def writeToTsv(filePath,table):
	newTable = []
	for line in table:
		newLine = []
		for item in line:
			if item.__class__ not in [str,unicode]:
				newLine.append(str(item))
			else:
				newLine.append(item)
		newTable.append("\t".join(newLine))
	writeStrToFile(filePath,"\n".join(newTable))

def readTsvToTable(tsvFile):
	table = []
	lines = readLinesFromFile(tsvFile)
	for line in lines:
		table.append(line.split("\t"))
	return table

def readLinesToTable(txtFile):
	table = []
	lines = readLinesFromFile(txtFile)
	for line in lines:
		table.append(line.split(" "))
	return table
	

def invertDict(mydict):
	inverted_dict = dict([[v, k] for k, v in mydict.items()])
	return inverted_dict
